import os.path
import html
import warnings
from typing import Callable, Union
from dateutil import parser
import openpyxl
import dateutil
from ..constants import MEDIA_TYPE_VIDEO, MEDIA_TYPE_GIF, MEDIA_TYPE_IMAGE
from ..exceptions import UserNotFound, UserProtected, ProtectedTweet
from ..utils import *


def deprecated(func):
    """

    This is a decorator which can be used to mark functions
    as deprecated. It will result in a warning being emitted
    when the function is used.

    """

    def new_func(*args, **kwargs):
        warnings.warn("Call to deprecated function {}.".format(func.__name__), category=DeprecationWarning)
        return func(*args, **kwargs)

    new_func.__name__ = func.__name__
    new_func.__doc__ = func.__doc__
    new_func.__dict__.update(func.__dict__)
    return new_func


class _TwType(dict):
    def __new__(cls, client, data, *args, **kwargs):
        if not data:
            return None

        return super().__new__(cls)

    def get_raw(self):
        return self._raw

    def __init_subclass__(cls, *args, **kwargs):
        super().__init_subclass__(*args, **kwargs)

        def new_init(self, *_args, init=cls.__init__, **_kwargs):
            init(self, *_args, **_kwargs) # noqa
            super().__init__()

            for k, v in vars(self).items():
                if not k.startswith("_"):
                    if isinstance(v, int):
                        self[k] = str(v)
                    else:
                        self[k] = v

        cls.__init__ = new_init


class Excel:
    def __init__(self, tweets, filename=None, append=False):
        self.tweets = tweets
        self.filename = filename
        self.author = self._get_author()
        self._append = append
        self._get_sheet()
        self.max_row = self._get_max_row()
        self._write_data()

    def _get_sheet(self):
        if self._append and self.filename:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
        else:
            self._append = False
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.create_sheet("tweets")

    def _get_author(self):
        for tweet in self.tweets:
            if hasattr(tweet, "author"):
                return tweet.author.username

        return ""

    def _get_max_row(self):
        if self._append:
            for row in reversed(self.worksheet.iter_rows(values_only=True)):
                if any(cell for cell in row):
                    return row[0].row

        self._set_headers()
        return 1

    def _set_headers(self):
        for index, value in enumerate(WORKBOOK_HEADERS, start=1):
            self.worksheet.cell(row=1, column=index).value = value

    def _write_tweet(self, tweet):
        self.worksheet[f'A{self.max_row + 1}'] = tweet.date.replace(tzinfo=None)
        self.worksheet[f'B{self.max_row + 1}'] = tweet.author.name
        self.worksheet[f'C{self.max_row + 1}'] = tweet.id
        self.worksheet[f'D{self.max_row + 1}'] = tweet.text
        self.worksheet[f'E{self.max_row + 1}'] = tweet.is_retweet
        self.worksheet[f'F{self.max_row + 1}'] = tweet.is_reply
        self.worksheet[f'G{self.max_row + 1}'] = tweet.language
        self.worksheet[f'H{self.max_row + 1}'] = tweet.likes
        self.worksheet[f'I{self.max_row + 1}'] = tweet.retweet_counts
        self.worksheet[f'J{self.max_row + 1}'] = tweet.source
        self.worksheet[f'K{self.max_row + 1}'] = iterable_to_string(tweet.media, ",", "direct_url")
        self.worksheet[f'L{self.max_row + 1}'] = iterable_to_string(tweet.user_mentions, ",", "screen_name")
        self.worksheet[f'M{self.max_row + 1}'] = iterable_to_string(tweet.urls, ",", "expanded_url")
        self.worksheet[f'N{self.max_row + 1}'] = iterable_to_string(tweet.hashtags, ",", "text")
        self.worksheet[f'O{self.max_row + 1}'] = iterable_to_string(tweet.symbols, ",", "text")
        self.max_row += 1

    def _write_data(self):
        for tweet in self.tweets:
            if isinstance(tweet, Tweet):
                self._write_tweet(tweet)
            elif isinstance(tweet, SelfThread):
                for _threadedTweet in tweet:
                    self._write_tweet(_threadedTweet)

        if not self.filename:
            self.filename = f"tweets-{self.author}.xlsx"

        try:
            self.workbook.remove(self.workbook["Sheet"])
        except ValueError:
            pass

        self.workbook.save(self.filename)


class EditControl(_TwType):
    def __init__(self, client, edit_control, parent, *args, **kwargs):
        self._client = client
        self._raw = edit_control['edit_control_initial'] if edit_control.get('edit_control_initial') else edit_control
        self._parent = parent
        self.tweet_ids = self._raw.get('edit_tweet_ids', [])
        self.edits_remaining = self._raw.get('edits_remaining', '0')
        self.is_edit_eligible = self._raw.get('is_edit_eligible', False)
        self.is_latest = str(self._parent.id) == str(self.tweet_ids[-1])
        self.latest_tweet_id = self.tweet_ids[-1]
        # self.history = [self.parent]

    @property
    def latest(self):
        if not self.is_latest:
            _latest = self._client.tweet_detail(self.tweet_ids[-1])
            return _latest
        return self._parent

    def __repr__(self):
        return "EditControl(parent={}, is_latest={})".format(
            self._parent, self.is_latest
        )


class Tweet(_TwType):
    def __init__(self, client, tweet, full_http_response=None, *args, **kwargs):  # noqa
        self._comments_cursor = None
        self._raw = tweet
        self._client = client
        self._full_http_response = full_http_response
        self._format_tweet()

    def __eq__(self, other):
        if isinstance(other, Tweet):
            return str(self.id) == str(other.id) and str(self.author.id) == str(other.author.id)
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return f"Tweet(id={self.id}, author={self.author}, created_on={self.created_on})"  # noqa

    def __iter__(self):
        if self.threads:  # noqa
            for thread in self.threads:  # noqa
                yield thread

    def like(self):
        return self._client.like_tweet(self.id)

    def unlike(self):
        return self._client.unlike_tweet(self.id)

    def retweet(self):
        return self._client.retweet_tweet(self.id)

    def translate(self, language):
        return self._client.translate_tweet(self.id, language)

    def delete(self):
        if self.author.id != self._client.me.id:
            return False

        return self._client.delete_tweet(self.id)

    def download_all_media(self, progress_callback=None):
        filenames = []
        for media in self.media:
            filenames.append(media.download(progress_callback=progress_callback))

        return filenames

    def get_threads(self):
        _threads = []
        if not self._full_http_response:
            return _threads

        instruction = find_objects(self._full_http_response, "type", "TimelineAddEntries")
        if not instruction:
            return _threads

        entries = instruction.get('entries', [])
        for entry in entries:

            if str(entry['entryId'].split("-")[0]) == "conversationthread":
                _thread = [i for i in entry['content']['items']]
                self_threads = [i for i in _thread if i['item']['itemContent'].get('tweetDisplayType') == "SelfThread"]

                if len(self_threads) == 0:
                    continue

                for _ in self_threads:
                    try:
                        parsed = Tweet(self._client, _, None)
                        _threads.append(parsed)
                    except:
                        pass

            elif str(entry['entryId'].split("-")[0]) == "tweet" and entry['content']['itemContent']['tweetDisplayType'] == "SelfThread":
                try:
                    parsed = Tweet(self._client, entry, None)
                    _threads.append(parsed)
                except:
                    pass
        return _threads

    def get_comments(self, pages=1, wait_time=2, cursor=None, get_hidden=False):
        return self._client.get_tweet_comments(self.id, pages, wait_time, cursor, get_hidden)

    def iter_comments(self, pages=1, wait_time=2, cursor=None, get_hidden=False):
        return self._client.iter_tweet_comments(self.id, pages, wait_time, cursor, get_hidden)

    def _check_if_protected(self):
        is_protected = is_tweet_protected(self._raw)

        if is_protected and is_protected.get('reason') == "Protected":
            raise ProtectedTweet(403, "TweetUnavailable", response=self._raw)
        elif is_protected and is_protected.get('reason') == "NsfwLoggedOut":
            raise AuthenticationRequired(401, "NsfwLoggedOut", response=self._raw, message="This Tweet is flagged as NSFW, make sure you are logged-in and brithday is updated in your account.")
        elif is_protected and is_protected.get('reason') == "Suspended":
            raise UserProtected(error_code="UserSuspended", response=self._raw, message="The Author of this Tweet is Suspended")
        elif is_protected and is_protected.get("tombstone"):
            error_message = is_protected.get("tombstone", {}).get("text", {}).get("text")
            if error_message:
                raise UserProtected(response=self._raw, message=error_message)
            else:
                raise UserProtected(response=self._raw)

    def _format_tweet(self):
        self._check_if_protected()
        self._tweet = find_objects(self._raw, "__typename", ["Tweet", "TweetWithVisibilityResults"], recursive=False)

        if self._tweet.get('tweet'):
            self._tweet = self._tweet['tweet']

        self._card = self._tweet.get('card')
        self._original_tweet = self._get_original_tweet()
        self.id = self._get_id()
        self.created_on = self.date = self._get_date()
        self.author = self._get_author()
        self.is_retweet = self._is_retweet()
        self.retweeted_tweet = self._get_retweeted_tweet()
        self.rich_text = self._get_rich_text()
        self.article = self._get_article()
        self.text = self.tweet_body = self._get_tweet_text()
        self.is_quoted = self._is_quoted()
        self.quoted_tweet = self._get_quoted_tweet()
        self.is_reply = self._is_reply()
        self.is_sensitive = self._is_sensitive()
        self.reply_counts = self._get_reply_counts()
        self.quote_counts = self._get_quote_counts()
        self.replied_to = None
        self.bookmark_count = self._get_bookmark_count()
        self.vibe = self._get_vibe()
        self.views = self._get_views()
        self.language = self._get_language()
        self.likes = self._get_likes()
        self.place = self._get_place()
        self.retweet_counts = self._get_retweet_counts()
        self.source = self._get_source()
        self.audio_space_id = self._get_audio_space()
        self.is_space = True if self.audio_space_id else False
        self.voice_info = None  # TODO
        self.media = self._get_tweet_media()
        self.pool = self._get_pool()
        self.user_mentions = self._get_tweet_mentions()
        self.urls = self._get_tweet_urls()
        self.has_moderated_replies = self._get_has_moderated_replies()
        self.hashtags = self._get_tweet_hashtags()
        self.symbols = self._get_tweet_symbols()
        self.community_note = self._get_community_note()
        self.community = self._get_community()
        self.url = self._get_url()
        self.edit_control = self._get_edit_control()
        self.has_newer_version = self._get_has_newer_version()
        self.broadcast = self._get_broadcast()
        self.threads = self.get_threads()
        self.is_liked = self._get_is_liked()
        self.is_retweeted = self._get_is_retweeted()
        self.can_reply = self._get_conversation_control()
        self.comments = []

    def _get_article(self):
        article_raw = self._tweet.get("article")
        if not article_raw:
            return None

        article_raw = article_raw.get("article_results", {}).get("result")
        return Article(self._client, article_raw)

    def _get_conversation_control(self):
        if not self._original_tweet.get('conversation_control') or self.author == self._client.me:
            return True

        conversation_control = self._original_tweet['conversation_control']
        control_policy = conversation_control.get('policy', '')

        if control_policy in ["Community", "Verified", "ByInvitation"]:
            actions = find_objects(self._raw, "limited_actions", None, none_value=[])
            if "limited_replies" in actions:
                return False
        return True

    def _get_is_liked(self):
        return self._original_tweet.get('favorited', False)

    def _get_is_retweeted(self):
        return self._original_tweet.get('retweeted', False)

    def _get_has_newer_version(self):
        if self.edit_control:
            return not self.edit_control.is_latest
        return False

    def get_latest(self):
        if self.edit_control:
            return self.edit_control.latest
        return self

    def _get_edit_control(self):
        edit_control = find_objects(self._raw, "edit_control", None, recursive=False)
        return EditControl(self._client, edit_control, self)

    def _get_broadcast(self):
        for url in self.urls:
            if "broadcast" in str(url):
                return Broadcast(self._client, self._card)

        return None

    def _get_url(self):
        return "https://twitter.com/{}/status/{}".format(
            self.author.username, self.id
        )

    def _get_original_tweet(self):
        tweet = self._tweet

        if tweet.get('tweet'):
            tweet = tweet['tweet']

        return tweet['legacy'] if tweet.get('legacy') else tweet

    def _get_has_moderated_replies(self):
        return self._tweet.get('hasModeratedReplies', False)

    def _get_pool(self):
        if not self._card or "poll" not in self._card.get('legacy', {}).get('name', ''):
            return None

        return Poll(self._client, self._card)

    def _get_audio_space(self):
        if not self._card or "audiospace" not in self._card.get('legacy', {}).get('name', ''):
            return None

        for value in self._card.get('legacy', {}).get('binding_values', []):
            if value['key'] == "id":
                return value['value']['string_value']

        return None

    def _get_community_note(self):
        if self._tweet.get("birdwatch_pivot"):
            text = self._tweet['birdwatch_pivot']['subtitle']['text']
            # entities = self._tweet['birdwatch_pivot']['subtitle'].get('entities', [])
            # for entity in entities:
            #     text = replace_between_indexes(text, entity['fromIndex'], entity['toIndex'], entity['ref']['url'])
            return text

        return None

    def _get_date(self):
        date = self._original_tweet.get("created_at")

        if date:
            return dateutil.parser.parse(date)

        return None

    def _get_id(self):
        return self._tweet.get('rest_id')

    def _get_community(self):
        return Community(self._client, self._tweet.get('community_results'))

    def _get_author(self):
        if self._tweet.get("core"):
            return User(self._client, self._tweet['core'])

        if self._tweet.get("author"):
            return User(self._client, self._tweet['author'])

        return None

    def _get_retweeted_tweet(self):
        if self.is_retweet and self._original_tweet.get("retweeted_status_result"):
            retweet = self._original_tweet['retweeted_status_result']['result']
            return Tweet(self._client, retweet)

        return None

    def _get_quoted_tweet(self):
        if not self.is_quoted:
            return None

        try:
            if self._tweet.get("quoted_status_result"):
                raw_tweet = self._tweet['quoted_status_result']['result']
                return Tweet(self._client, raw_tweet)

            if self._original_tweet.get('retweeted_status_result'):
                raw_tweet = self._original_tweet['retweeted_status_result']['result']['quoted_status_result']['result']
                return Tweet(self._client, raw_tweet)

            return None
        except:
            return None

    def _get_vibe(self):
        if self._tweet.get("vibe"):
            vibeImage = self._tweet['vibe']['imgDescription']
            vibeText = self._tweet['vibe']['text']
            return f"{vibeImage} {vibeText}"

        return ""

    def _get_views(self):
        if self._tweet.get("views"):
            return self._tweet['views'].get('count', 'Unavailable')

        return 0

    def get_reply_to(self):
        if not self.is_reply:
            return None

        tweet_id = self._original_tweet['in_reply_to_status_id_str']
        if not self._full_http_response:
            response = self._client.http.get_tweet_detail(tweet_id)
        else:
            response = self._full_http_response

        try:
            entries = find_objects(response, "entries", None, none_value=[])
            for entry in entries:
                if str(entry['entryId']).split("-")[0] == "tweet" and str(
                        entry['content']['itemContent']['tweet_results']['result']['rest_id']) == str(tweet_id):
                    raw_tweet = entry['content']['itemContent']['tweet_results']['result']
                    return Tweet(self._client, raw_tweet)
        except:
            pass

        return None

    def _is_sensitive(self):
        return self._original_tweet.get("possibly_sensitive", False)

    def _get_reply_counts(self):
        return self._original_tweet.get("reply_count", 0)

    def _get_quote_counts(self):
        return self._original_tweet.get("quote_count", 0)

    def _is_retweet(self):
        if self._original_tweet.get('retweeted'):
            return self._original_tweet['retweeted']

        if str(self._original_tweet.get('full_text', "")).startswith("RT"):
            return True

        return False

    def _is_reply(self):
        tweet_keys = list(self._original_tweet.keys())
        required_keys = ["in_reply_to_status_id_str", "in_reply_to_user_id_str", "in_reply_to_screen_name"]
        return all(x in tweet_keys for x in required_keys)

    def _is_quoted(self):
        if self._original_tweet.get("is_quote_status"):
            return True

        return False

    def _get_language(self):
        return self._original_tweet.get('lang', "")

    def _get_likes(self):
        return self._original_tweet.get("favorite_count", 0)

    def _get_place(self):
        if self._original_tweet.get('place'):
            return Place(self._client, self._original_tweet['place'])

        return None

    def _get_retweet_counts(self):
        return self._original_tweet.get('retweet_count', 0)

    def _get_source(self):
        if self._tweet.get('source'):
            return str(self._tweet['source']).split(">")[1].split("<")[0]

        return ""

    def _get_rich_text(self):
        note_tweet = self._tweet.get('note_tweet')

        if not note_tweet:
            return None

        return RichText(self._client, note_tweet, self)

    def _get_tweet_text(self):
        if self.is_retweet and self._original_tweet.get("retweeted_status_result"):
            return self.retweeted_tweet.text

        if self.rich_text:
            return self.rich_text.text

        if self.article:
            return self.article.text

        if self._original_tweet.get('full_text'):
            return html.unescape(self._original_tweet['full_text'])

        return ""

    def _get_tweet_media(self):
        if self.is_retweet and self.retweeted_tweet:
            return self.retweeted_tweet.media

        return [Media(self._client, media) for media in
                self._original_tweet.get("extended_entities", {}).get("media", [])]

    def _get_tweet_mentions(self):
        users = [ShortUser(self._client, user) for user in
                 self._original_tweet.get("entities", {}).get("user_mentions", [])]

        if self.rich_text:
            for user in self.rich_text.user_mentions:
                if user not in users:
                    users.append(user)
        return users

    def _get_bookmark_count(self):
        return self._original_tweet.get("bookmark_count", None)

    def _get_tweet_urls(self):
        urls = self._original_tweet.get('entities', {}).get('urls', [])
        urls = [URL(self._client, url) for url in urls]

        if self.rich_text:
            urls.extend(self.rich_text.urls)

        return urls

    def _get_tweet_hashtags(self):
        hashtags = self._original_tweet.get('entities', {}).get('hashtags', [])
        hashtags = [Hashtag(self._client, hashtag) for hashtag in hashtags]

        if self.rich_text:
            hashtags.extend(self.rich_text.hashtags)

        return hashtags

    def _get_tweet_symbols(self):
        symbols = self._original_tweet.get('entities', {}).get('symbols', [])
        symbols = [Symbol(self._client, symbol) for symbol in symbols]

        if self.rich_text:
            symbols.extend(self.rich_text.symbols)

        return symbols


class Symbol(_TwType):
    def __init__(self, client, symbol, *args, **kwargs):
        self._client = client
        self._raw = symbol
        self.indices = self._raw['indices']
        self.text = self._raw['text']

    def __repr__(self):
        return "Symbol(text={})".format(self.text)


class URL(_TwType):
    def __init__(self, client, url, *args, **kwargs):
        self._client = client
        self._raw = url
        self.display_url = self._raw.get('display_url')
        self.expanded_url = self._raw.get('expanded_url')
        self.url = self._raw.get('url')
        self.indices = self._raw.get('indices')

    def __str__(self):
        return self.expanded_url

    def __repr__(self):
        return "URL(expanded_url={})".format(self.expanded_url)


class Hashtag(_TwType):
    def __init__(self, client, hashtag, *args, **kwargs):
        self._client = client
        self._raw = hashtag
        self.indices = self._raw.get('indices')
        self.text = self._raw.get('text')

    def __str__(self):
        return self.text

    def __repr__(self):
        return "Hashtag(text={})".format(self.text)


class RichText(_TwType):
    HTML_TAGS = {
        "Bold": "b",
        "Italic": "em"
    }

    def __init__(self, client, data, tweet, *args, **kwargs):
        self._client = client
        self._raw = data
        self._tweet = tweet
        self._note = self._raw.get('note_tweet_results', {}).get('result', {})
        self._entities = self._note.get('entity_set', {})
        self.tags = self._get_tags()
        self.id = self._get_id()
        self.text = self._get_text()
        self.hashtags = self._get_hashtags()
        self.urls = self._get_urls()
        self.symbols = self._get_symbols()
        self.user_mentions = self._get_mentions()
        self.media = self._get_media()

    def __repr__(self):
        return "RichText(id={})".format(self.id)

    def __eq__(self, other):
        if isinstance(other, RichText):
            return self.id == other.id
        elif isinstance(other, str):
            return str(self.id) == str(other)

        return False

    def _get_id(self):
        return self._note.get('id')

    def _get_text(self):
        return self._note.get('text', '')

    def _get_hashtags(self):
        return [Hashtag(self._client, i) for i in self._entities.get('hashtags', [])]

    def _get_urls(self):
        return [URL(self._client, i) for i in self._entities.get('urls', [])]

    def _get_symbols(self):
        return [Symbol(self._client, symbol) for symbol in self._entities.get('symbols', [])]

    def _get_tags(self):
        return [RichTag(self._client, i) for i in self._note.get('richtext', {}).get('richtext_tags', [])]

    def _get_media(self):
        return [i for i in self._note.get('media', {}).get('inline_media', [])]

    def _get_mentions(self):
        return [ShortUser(self._client, i) for i in self._entities.get('user_mentions', [])]

    def get_html(self):
        tags = self.tags
        tags.extend(self.user_mentions)
        tags.extend(self.media)
        ordered_tags = sorted(tags, key=lambda x: x.from_index if hasattr(x, "from_index") else x['index'],
                              reverse=True)
        thisHtml = self.text

        for tag in ordered_tags:
            text = self.text[tag.from_index:tag.to_index] if hasattr(tag, "from_index") else self.text[tag['index']]
            if isinstance(tag, RichTag) and hasattr(tag, "types"):
                for _type in tag.types:
                    tag_name = self.HTML_TAGS.get(_type)
                    new_text = "<{tag_name}>{text}</{tag_name}>".format(tag_name=tag_name, text=text)
                    thisHtml = replace_between_indexes(thisHtml, tag.from_index, tag.to_index, new_text)
            elif isinstance(tag, ShortUser):
                new_text = "<a href='{}'>@{}</a>".format(tag.url, tag.username)
                thisHtml = replace_between_indexes(thisHtml, tag.from_index, tag.to_index, new_text)
            elif tag.get('media_id'):
                for media in self._tweet.media:
                    if media == tag['media_id']:
                        new_text = "<img src='{}'><br>".format(media.direct_url)
                        thisHtml = replace_between_indexes(thisHtml, tag['index'], tag['index'], new_text)
                        break

        return f"<pre>{thisHtml}</pre>"


class RichTag(_TwType):
    def __init__(self, client, data, *args, **kwargs):
        self._client = client
        self._raw = data
        self.from_index = self._raw.get('from_index')
        self.to_index = self._raw.get('to_index')
        self.types = self._raw.get('richtext_types', [])

    def __repr__(self):
        return "RichTag(from_index={}, to_index={}, types={})".format(
            self.from_index, self.to_index, self.types
        )


class SelfThread(_TwType):
    def __init__(self, client, conversation_tweet, *args, **kwargs):
        self._client = client
        self._raw = conversation_tweet
        self.tweets = []
        self.all_tweets_id = self._get_all_tweet_ids()
        self._format_tweet()

    def _format_tweet(self):
        for item in self._raw['content']['items']:
            try:
                parsed = Tweet(self._client, item, None)
                self.tweets.append(parsed)
            except:
                pass

    def __iter__(self):
        if self.tweets:
            for tweet in self.tweets:  # noqa
                yield tweet

    def expand(self):
        tweet = self._client.tweet_detail(self.tweets[0].id)
        self.tweets = tweet.threads

    def _get_all_tweet_ids(self):
        return find_objects(self._raw, 'allTweetIds', None, none_value=[], recursive=False)

    def __repr__(self):
        return "SelfThread(tweets={}, all_tweets={})".format(
            len(self.tweets), len(self.all_tweets_id)
        )


class ConversationThread(_TwType):
    def __init__(self, client, parent_tweet, thread_tweets, *args, **kwargs):
        self._client = client
        self.tweets = []
        self.parent = parent_tweet
        self.cursor = None
        self._threads = thread_tweets
        self._format_threads()

    def __repr__(self):
        return "ConversationThread(parent={}, tweets={})".format(
            self.parent, len(self.tweets)
        )

    def __iter__(self):
        if self.tweets:
            for tweet in self.tweets:  # noqa
                yield tweet

    def _format_threads(self):
        for thread in self._threads:
            entry_type = str(thread['entryId']).split("-")[-2].lower()
            if entry_type == "tweet":
                self.tweets.append(Tweet(self._client, thread, None))
            elif entry_type == "showmore":
                self.cursor = thread['item']['itemContent']['value']

    def expand(self):
        if not self.cursor:
            return self.tweets

        response = self._client.http.get_tweet_detail(self.parent.id, self.cursor)
        moduleItems = find_objects(response, "moduleItems", None)

        if not moduleItems or len(moduleItems) == 0:
            return self.tweets

        for item in moduleItems:
            tweet = find_objects(item, "__typename", ["Tweet", "TweetWithVisibilityResults"], recursive=False)
            if tweet:
                self.tweets.append(Tweet(self._client, tweet, None))

        return self.tweets


class Media(_TwType):
    def __init__(self, client, media_dict, *args, **kwargs):
        self._raw = media_dict
        self._client = client
        self.display_url = self._raw.get("display_url")
        self.expanded_url = self._raw.get("expanded_url")
        self.id = self._raw.get("id_str")
        self.alt_text = self._raw.get("ext_alt_text")
        self.indices = self._raw.get("indices")
        self.audio_only = self._raw.get('audio_only')
        self.media_url_https = self.direct_url = self._get_direct_url()
        self.type = self._raw.get("type")
        self.url = self._raw.get("url")
        self.features = self._raw.get("features")
        self.media_key = self._raw.get("media_key")
        self.mediaStats = self._raw.get("mediaStats")
        self.sizes = [MediaSize(self._client, v, k) for k, v in self._raw.get("sizes", {}).items() if
                      self._raw.get('sizes')]
        self.original_info = self._raw.get("original_info")
        self.file_format = self._get_file_format()
        self.source_user = self._get_source_user()
        self.streams = []

        if self.type in (MEDIA_TYPE_VIDEO, MEDIA_TYPE_GIF):
            self._parse_video_streams()

    def __eq__(self, other):
        if isinstance(other, Media):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def _get_direct_url(self):
        url = self._raw.get("media_url_https")
        return url

    def _get_file_format(self):
        filename = os.path.basename(self.media_url_https).split("?")[0]
        return filename.split(".")[-1] if self.type == MEDIA_TYPE_IMAGE else "mp4"

    def _parse_video_streams(self):
        videoDict = self._raw.get("video_info")

        if not videoDict:
            return

        for i in videoDict.get("variants"):
            if not i.get("content_type").split("/")[-1] == "x-mpegURL":
                self.streams.append(Stream(self._client, i, videoDict.get("duration_millis", 0), videoDict.get("aspect_ratio")))

    def best_stream(self):
        if self.type == MEDIA_TYPE_IMAGE:
            return self
        elif self.type == MEDIA_TYPE_VIDEO:
            _res = [eval(stream.res) for stream in self.streams if stream.res]
            max_res = max(_res)
            for stream in self.streams:
                if eval(stream.res) == max_res:
                    file_format = stream.content_type.split("/")[-1]
                    if not file_format == "x-mpegURL":
                        return stream
        elif self.type == MEDIA_TYPE_GIF:
            for stream in self.streams:
                file_format = stream.content_type.split("/")[-1]
                if not file_format == "x-mpegURL":
                    return stream
        return None

    def __repr__(self):
        return f"Media(id={self.id}, type={self.type}, audio_only={self.audio_only})"

    def download(self, filename: str = None, progress_callback: Callable[[str, int, int], None] = None):
        url = self.best_stream().direct_url

        if not url:
            raise ValueError("No Media Download URL found")

        return self._client.http.download_media(url, filename, progress_callback)

    def _get_source_user(self):
        source_user = find_objects(self._raw, "source_user", None, recursive=False)
        if not source_user:
            return None

        user = find_objects(source_user, "__typename", "User", recursive=False)
        try:
            user = User(self._client, user)
        except:
            user = None
        return user


class Stream(_TwType):
    def __init__(self, client, video_dict, length, ratio, *args, **kwargs):
        self._raw = video_dict
        self._client = client
        self.bitrate = self._raw.get("bitrate")
        self.content_type = self._raw.get("content_type")
        self.url = self.direct_url = self._raw.get("url")
        self.length = length
        self.aspect_ratio = ratio
        self.res = self._get_resolution()

    def _get_resolution(self):
        result = re.findall("/(\d+)x(\d+)/", self.url)
        if result:
            result = result[0]
            return f"{result[0]}*{result[1]}"

        return None

    def __repr__(self):
        return f"Stream(content_type={self.content_type}, length={self.length}, bitrate={self.bitrate}, res={self.res})"

    def download(self, filename: str = None, progress_callback: Callable[[str, int, int], None] = None):
        return self._client.http.download_media(self.url, filename, progress_callback)


class MediaSize(_TwType):
    def __init__(self, client, data, name, *args, **kwargs):
        self._client = client
        self._raw = data
        self.name = self['name'] = name
        self.width = self['width'] = self._raw.get('w')
        self.height = self['height'] = self._raw.get('h')
        self.resize = self['resize'] = self._raw.get('resize')

    def __repr__(self):
        return "MediaSize(name={}, width={}, height={}, resize={})".format(
            self.name, self.width, self.height, self.resize
        )


class ShortUser(_TwType):
    def __init__(self, client, user_dict, *args, **kwargs):
        self._client = client
        self.__raw = user_dict
        self._indices = self.__raw.get("indices")
        self.id = self.__raw.get("id_str")
        self.name = self.__raw.get("name")
        self.screen_name = self.username = self.__raw.get("screen_name")
        self.url = "https://twitter.com/{}".format(self.username)
        self.from_index = self._indices[0] if self._indices else None
        self.to_index = self._indices[1] if self._indices else None

    def get_full_user(self):
        return self._client.get_user_info(self.id)

    def __eq__(self, other):
        if isinstance(other, ShortUser):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return f"ShortUser(id={self.id}, name={self.name})"


class Trends(_TwType):
    def __init__(self, client, trend_item, *args, **kwargs):
        self._raw = trend_item
        self._trend = find_objects(self._raw, "trend", None)
        self.name = self._get_name()
        self.url = self._get_url()
        self.tweet_count = self._get_count()

    def _get_name(self):
        return self._trend.get('name')

    def _get_url(self):
        url = self._trend['url'].get('url')
        url = url.replace("twitter://", "https://twitter.com/").replace("query", "q")
        return url

    def _get_count(self):
        return self._trend.get('trendMetadata', {}).get('metaDescription')

    def __repr__(self):
        return f"Trends(name={self.name})"


class Broadcast(_TwType):
    def __init__(self, client, broadcast, *args, **kwargs):
        self._client = client
        self._raw = broadcast
        self._broadcast = self._raw.get('legacy', {})
        self._parsed = self._parse_keys()
        self.url = self._get_url()
        self.width = self._get_width()
        self.state = self._get_state()
        self.title = self._get_title()
        self.source = self._get_source()
        self.id = self._get_id()
        self.user_id = self.broadcaster_id = self._get_broadcaster_id()
        self.height = self._get_height()
        self.username = self.broadcaster_username = self._get_broadcaster_username()
        self.media_key = self._get_media_key()
        self.broadcaster_name = self._get_broadcaster_display_name()
        self.media_id = self._get_media_id()
        self.thumbnail_large = self._get_image("broadcast_thumbnail_large")
        self.thumbnail = self._get_image("broadcast_thumbnail")
        self.thumbnail_x_large = self._get_image("broadcast_thumbnail_x_large")
        self.thumbnail_original = self._get_image("broadcast_thumbnail_original")
        self.thumbnail_small = self._get_image("broadcast_thumbnail_small")

    def _get_url(self):
        return self._parsed.get('broadcast_url', {}).get('string_value')

    def _get_width(self):
        return self._parsed.get('broadcast_width', {}).get('string_value')

    def _get_state(self):
        return self._parsed.get('broadcast_state', {}).get('string_value')

    def _get_title(self):
        return self._parsed.get('broadcast_title', {}).get('string_value')

    def _get_source(self):
        return self._parsed.get('broadcast_source', {}).get('string_value')

    def _get_id(self):
        return self._parsed.get('broadcast_id', {}).get('string_value')

    def _get_broadcaster_id(self):
        return self._parsed.get('broadcaster_twitter_id', {}).get('string_value')

    def _get_height(self):
        return self._parsed.get('broadcast_height', {}).get('string_value')

    def _get_broadcaster_username(self):
        return self._parsed.get('broadcaster_username', {}).get('string_value')

    def _get_media_key(self):
        return self._parsed.get('broadcast_media_key', {}).get('string_value')

    def _get_broadcaster_display_name(self):
        return self._parsed.get('broadcaster_display_name', {}).get('string_value')

    def _get_media_id(self):
        return self._parsed.get('broadcast_media_id', {}).get('string_value')

    def _get_image(self, key):
        return self._parsed.get(key, {}).get('value', {}).get('image_value')

    def _parse_keys(self):
        parsed = {}
        if isinstance(self._broadcast['binding_values'], list):
            for value in self._broadcast['binding_values']:
                key, value_ = value['key'], value['value']
                parsed[key] = value_
        elif isinstance(self._broadcast['binding_values'], dict):
            parsed = self._broadcast['binding_values']
        return parsed

    def __repr__(self):
        return "Broadcast(id={}, title={}, state={}, broadcaster_username={})".format(
            self.id, self.title, self.state, self.broadcaster_username
        )


class Poll(_TwType):
    CHOICE_LABEL_REGEX = r"choice\d+_label"
    CHOICE_COUNT_REGEX = r"choice\d+_count"
    CHOICE_COUNT_FORMAT = "choice{}_count"

    def __init__(self, client, card, *args, **kwargs):
        self._client = client
        self._raw = card
        self._pool = self._raw['legacy']
        self._parsed = self._parse_keys()
        self.id = self._get_id()
        self.name = self._get_name()
        self.choices = self._get_choices()
        self.end_time = self._get_end_time()
        self.last_updated_time = self._get_last_updated_time()
        self.is_final = self._get_is_final()
        self.duration = self._get_duration()
        self.selected_choice = self._get_selected_choice()
        self.user_ref = self._get_user_ref()

    def __repr__(self):
        return "Pool(id={}, end_time={}, duration={} minutes, is_final={}, choices={})".format(
            self.id, self.end_time, self.duration, self.is_final, self.choices
        )

    def _parse_keys(self):
        parsed = {}
        if isinstance(self._pool['binding_values'], list):
            for value in self._pool['binding_values']:
                key, value_ = value['key'], value['value']
                parsed[key] = value_
        elif isinstance(self._pool['binding_values'], dict):
            parsed = self._pool['binding_values']
        return parsed

    def _get_name(self):
        return self._pool.get('name')

    def _get_id(self):
        return self._raw.get('rest_id') or self._pool.get('url', '').replace("\\", "")

    def _get_choices(self):
        results = []

        for key, value in self._parsed.items():
            if re.match(self.CHOICE_LABEL_REGEX, key):
                number = re.findall("\d+", key)[0]
                string_value = value['string_value']
                choice_count = self._parsed.get(self.CHOICE_COUNT_FORMAT.format(number), {}).get('string_value', "0")

                results.append(Choice(number, self.id, self.name, string_value, choice_count))
        return results

    def _get_selected_choice(self):
        choice = self._parsed.get('selected_choice', None)

        if not choice:
            return None

        for _choice in self.choices:
            if _choice.key == choice['string_value']:
                return _choice

        return None

    def _get_user_ref(self):
        return [User(self._client, user) for user in self._pool.get('user_refs_results', [])]

    def _get_end_time(self):
        return parse_time(self._parsed.get('end_datetime_utc', {}).get('string_value'))

    def _get_last_updated_time(self):
        return parse_time(self._parsed.get('last_updated_datetime_utc', {}).get('string_value'))

    def _get_duration(self):
        return self._parsed.get('duration_minutes', {}).get('string_value', "0")

    def _get_is_final(self):
        return self._parsed.get('counts_are_final', {}).get('boolean_value', False)


class Choice(dict):
    def __init__(self, key, pool_id, pool_name, choice_value, choice_count):
        self.key = key
        self.name = pool_name
        self.id = pool_id
        self.value = choice_value
        self.counts = choice_count

    def __repr__(self):
        return f"Choice(key={self.key}, value={self.value}, counts={self.counts})"


class Place(_TwType):
    def __init__(self, client, place_dict, *args, **kwargs):
        self._client = client
        self._raw = place_dict.get('place') if place_dict.get('place') else place_dict
        self.id = self._raw.get("id")
        self.country = self._raw.get("country")
        self.country_code = self._raw.get("country_code")
        self.full_name = self._raw.get("full_name")
        self.name = self._raw.get("name")
        self.type = self._raw.get("place_type")
        self.attributes = self._raw.get("attributes")
        self.url = self._raw.get("url")
        self.coordinates = self.parse_coordinates()
        self.centroid = self._get_centroid()

    def _get_centroid(self):
        if not self._raw.get('centroid'):
            return None

        return Coordinates(*self._raw['centroid'])

    def parse_coordinates(self):
        results = []

        if not self._raw.get("bounding_box"):
            return results

        for i in self._raw['bounding_box'].get("coordinates"):
            for p in i:
                coordinates = p[:2]

                if coordinates not in results:
                    results.append(coordinates)

        return [Coordinates(*i) for i in results]

    def __repr__(self):
        return f"Place(id={self.id}, name={self.name}, country={self.country, self.country_code}, full_name={self.full_name})"


class Coordinates(dict):
    def __init__(self, latitude, longitude):
        super().__init__()
        self.latitude = latitude
        self.longitude = longitude

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __repr__(self):
        return f"Coordinates(latitude={self.latitude}, longitude={self.longitude})"


class User(_TwType):
    def __init__(self, client, user_data, *args, **kwargs):
        self._raw = user_data
        self._client = client
        self._user = find_objects(self._raw, "__typename", "User", recursive=False)

        if not self._user:
            if find_objects(self._raw, "__typename", "UserUnavailable", recursive=False):
                message = find_objects(self._raw, "message", None, recursive=False,none_value=None)
                raise UserProtected(response=user_data, message=message)
            else:
                raise UserNotFound(response=user_data)

        self._original_user = self._user['legacy'] if self._user.get('legacy') else self._user
        self._social_context = self._user.get('social_context', {})
        self.id = self.rest_id = self.get_id()
        self.created_at = self.date = self.get_created_at()
        self.entities = self._get_key("entities")
        self.description = self.bio = self._get_key("description")
        self.fast_followers_count = self._get_key("fast_followers_count", default=0)
        self.favourites_count = self._get_key("favourites_count", default=0)
        self.followers_count = self._get_key("followers_count", default=0)
        self.friends_count = self._get_key("friends_count", default=0)
        self.has_custom_timelines = self._get_key("has_custom_timelines", default=False)
        self.is_translator = self._get_key("is_translator", default=False)
        self.listed_count = self._get_key("listed_count", default=0)
        self.location = self._get_key("location")
        self.media_count = self._get_key("media_count", default=0)
        self.name = self._get_key("name")
        self.normal_followers_count = self._get_key("normal_followers_count", default=0)
        self.subscriptions_count = self._get_key("creator_subscriptions_count", default=0)
        self.profile_banner_url = self._get_key("profile_banner_url")
        self.profile_image_url_https = self._get_key("profile_image_url_https")
        self.profile_interstitial_type = self._get_key("profile_interstitial_type")
        self.protected = self._get_key("protected", default=False)
        self.screen_name = self.username = self._get_key("screen_name")
        self.statuses_count = self._get_key("statuses_count", default=0)
        self.translator_type = self._get_key("translator_type")
        self.verified = self._get_verified()
        self.can_dm = self._get_key("can_dm")
        self.following = self._get_key("following", False)
        self.followed_by = self._get_key("followed_by", False)
        self.community_role = self._get_key("community_role", None)
        self.notifications_enabled = self.notifications = self._get_key("notifications", False)
        # self.verified_type = self._get_key("verified_type")
        self.possibly_sensitive = self._get_key("possibly_sensitive", default=False)
        self.pinned_tweets = self._get_key("pinned_tweet_ids_str")
        self.profile_url = "https://twitter.com/{}".format(self.screen_name)
        self.is_blocked = self._get_is_blocked()
        self.blocked_by = self.has_blocked_me = self._get_blocked_by()

    def __eq__(self, other):
        if isinstance(other, (User, ShortUser)):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return "User(id={}, username={}, name={}, verified={})".format(
            self.id, self.username, self.name, self.verified
        )

    def follow(self):
        return self._client.follow_user(self.id)

    def unfollow(self):
        return self._client.unfollow_user(self.id)
    
    def block(self):
        return self._client.block_user(self.id)
    
    def unblock(self):
        return self._client.unblock_user(self.id)

    def enable_notifications(self):
        if not self.notifications:
            return self._client.enable_user_notification(self.id)

        return True

    def disable_notifications(self):
        if self.notifications:
            return self._client.disable_user_notification(self.id)

        return True

    def add_to_list(self, list_id):
        return self._client.add_list_member(list_id, self.id)

    def remove_from_list(self, list_id):
        return self._client.remove_list_member(list_id, self.id)

    def _get_is_blocked(self):
        return self._original_user.get('blocking', False)

    def _get_blocked_by(self):
        return self._original_user.get('blocked_by', False)

    def _get_verified(self):
        verified = self._get_key("verified", False)
        if verified is False:
            verified = self._get_key("is_blue_verified", False)

        if verified is False:
            verified = self._get_key("ext_is_blue_verified", False)

        return False if verified in (None, False) else True

    def get_id(self):
        raw_id = self._user.get("id")

        if not raw_id:
            raw_id = self._user.get('rest_id')

        if not str(raw_id).isdigit():
            raw_id = decodeBase64(raw_id).split(":")[-1]

        return int(raw_id)

    def get_created_at(self):
        return parse_time(self._original_user.get('created_at'))

    def _get_key(self, key, default=None):
        keyValue = default

        if self._user.get(key):
            keyValue = self._user[key]

        if self._original_user.get(key):
            keyValue = self._original_user[key]

        if self._social_context.get(key):
            keyValue = self._social_context[key]

        if str(keyValue).isdigit():
            keyValue = int(keyValue)

        return keyValue


class PeriScopeUser(_TwType):
    def __init__(self, client, user_data, *args, **kwargs):
        self._client = client
        self._raw = user_data
        self.id = self._raw.get('periscope_user_id')
        self.twitter_screen_name = self.username = self._raw.get('twitter_screen_name')
        self.display_name = self.name = self._raw.get('display_name')
        self.is_verified = self._raw.get('is_verified')
        self.twitter_id = self._raw['user_results'].get('rest_id')


class AudioSpace(_TwType):
    def __init__(self, client, audio_space, *args, **kwargs):
        self._raw = audio_space
        self._client = client
        self._space = find_objects(self._raw, "audioSpace", None, recursive=False)
        self._meta_data = find_objects(self._space, "metadata", None, recursive=False, none_value={})
        self._participants = find_objects(self._raw, "participants", None, recursive=False)
        self.id = self._meta_data.get('rest_id')
        self.state = self._meta_data.get('state')
        self.title = self._meta_data.get('title')
        self.media_key = self._meta_data.get('media_key')
        self.created_at = parse_time(self._meta_data.get('created_at'))
        self.started_at = parse_time(self._meta_data.get('started_at'))
        self.ended_at = parse_time(self._meta_data.get('ended_at'))
        self.updated_at = parse_time(self._meta_data.get('updated_at'))
        self.creator = User(self._client, self._meta_data.get('creator_results'))
        self.total_live_listeners = self._meta_data.get('total_live_listeners')
        self.total_replay_watched = self._meta_data.get('total_replay_watched')
        self.disallow_join = self._meta_data.get('disallow_join')
        self.is_employee_only = self._meta_data.get('is_employee_only')
        self.is_locked = self._meta_data.get('is_locked')
        self.is_muted = self._meta_data.get('is_muted')
        self.tweet = Tweet(self._client, self._meta_data.get('tweet_results'))
        self.admins = self._get_participants('admins')
        self.speakers = self._get_participants('speakers')

    def _get_participants(self, participant):
        return [PeriScopeUser(self._client, user) for user in self._participants[participant]]

    def get_stream_link(self):
        return self._client.http.get_audio_stream(self.media_key)

    def __repr__(self):
        return "AudioSpace(id={}, title={}, state={}, tweet={})".format(
            self.id, self.title, self.state, self.tweet
        )


class Community(_TwType):
    def __init__(self, client, data, *args, **kwargs):
        self._raw = data
        self._client = client
        self._community = find_objects(self._raw, "__typename", "Community", recursive=False)
        self.id = self._get_id()
        self.date = self.created_at = self._get_date()
        self.description = self._get_description()
        self.name = self._get_name()
        self.role = self._get_role()
        self.member_count = self._get_member_count()
        self.moderator_count = self._get_moderator_count()
        self.admin = self._get_admin()
        self.creator = self._get_creator()
        self.rules = self._get_rules()

    def __repr__(self):
        return "Community(id={}, name={}, role={}, admin={})".format(
            self.id, self.name, self.role, self.admin
        )

    def _get_id(self):
        return self._community.get('id_str')

    def _get_date(self):
        return parse_time(self._community.get('created_at'))

    def _get_description(self):
        return self._community.get('description')

    def _get_name(self):
        return self._community.get('name')

    def _get_member_count(self):
        return self._community.get('member_count')

    def _get_moderator_count(self):
        return self._community.get('moderator_count')

    def _get_admin(self):
        return User(self._client, self._community['admin_results'])

    def _get_creator(self):
        return User(self._client, self._community['creator_results'])

    def _get_rules(self):
        return [rule['name'] for rule in self._community.get('rules', [])]

    def _get_role(self):
        return self._community.get('role')


class List(_TwType):
    def __init__(self, client, list_data, *args, **kwargs):
        self._raw = list_data
        self._client = client
        self._list = self._get_list()
        self.id = self._get_id()
        self.name = self._get_name()
        self.created_at = self.date = self._get_date()
        self.description = self._get_description()
        self.is_member = self._get_is_member()
        self.member_count = self._get_member_count()
        self.subscriber_count = self._get_subscriber_count()
        self.admin = self._get_admin()
        self.mode = self._get_mode()

    def __eq__(self, other):
        if isinstance(other, List):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return "List(id={}, name={}, admin={}, subscribers={})".format(
            self.id, self.name, self.admin, self.subscriber_count
        )

    def _get_list(self):
        if self._raw.get('entryId'):
            self._raw = find_objects(self._raw, "__typename", "TimelineTwitterList", {})

        if self._raw.get('list'):
            return self._raw['list']

        return self._raw

    def _get_id(self):
        if self._list.get('id_str'):
            return int(self._list['id_str'])

        if self._list.get("id"):
            _id = decodeBase64(self._list['id'])
            return int(str(_id).replace("List:", ""))

        return None

    def _get_name(self):
        return self._list.get('name')

    def _get_date(self):
        return parse_time(self._list.get('created_at'))

    def _get_description(self):
        return self._list.get("description")

    def _get_is_member(self):
        return self._list.get('is_member', False)

    def _get_member_count(self):
        return self._list.get('member_count', 0)

    def _get_subscriber_count(self):
        return self._list.get('subscriber_count', 0)

    def _get_mode(self):
        return self._list.get('mode')

    def _get_admin(self):
        if not self._list.get('user_results'):
            return None

        return User(self._client, self._list['user_results'])


class Gif(_TwType):
    def __init__(self, client, gif, *args, **kwargs):
        self._client = client
        self._raw = gif
        self.provider = self._raw.get('provider', {}).get('name')
        self.id = self._raw.get('id')
        self.alt_text = self._raw.get('alt_text')
        self.url = self._raw.get('original_image', {}).get('url')

    def __repr__(self):
        return "Gif(id={}, provider={}, alt_text={})".format(
            self.id, self.provider, self.alt_text
        )


class Topic(_TwType):
    def __init__(self, client, topic):
        self._client = client
        self._raw = topic
        self.original_topic = find_objects(self._raw, "__typename", ["TopicPageHeader", "TimelineTopic"], recursive=False)

        if not self.original_topic:
            raise ValueError("Topic Not Found")

        self.original_topic = self.original_topic['topic'] if self.original_topic.get('topic') else self.original_topic

        self.id = self.original_topic.get('topic_id')
        self.description = self.original_topic.get('description')
        self.name = self.original_topic.get('name')
        self.is_following = self.original_topic.get('following')
        self.icon_url = self.original_topic.get('icon_url')
        self.is_not_interested = self.original_topic.get('not_interested')

    def __repr__(self):
        return "Topic(id={}, name={})".format(self.id, self.name)


class TweetTranslate(_TwType):
    def __init__(self, client, translate, *args, **kwargs):
        self._client = client
        self._raw = translate
        self.id = self._raw.get('id')
        self.translation = self.text = self._raw.get('translation')
        self.source_language = self._raw.get('sourceLanguage')
        self.destination_language = self._raw.get('destinationLanguage')
        self.localized_source_language = self._raw.get('localizedSourceLanguage')

    def __repr__(self):
        return "TweetTranslate(id={}, source_language={})".format(self.id, self.source_language)


class TweetAnalytics(_TwType):
    def __init__(self, client, analytics):
        self._raw = analytics
        self._client = client
        self._tweet = find_objects(self._raw, "__typename", "Tweet", none_value={})
        self.expands = self._get_metric("DetailExpands")
        self.engagements = self._get_metric("Engagements")
        self.follows = self._get_metric("Follows")
        self.impressions = self._get_metric("Impressions")
        self.link_clicks = self._get_metric("LinkClicks")
        self.profile_visits = self._get_metric("ProfileVisits")
        self.cost_per_follower = self._get_metric("CostPerFollower")

    def _get_metric(self, metric_key):
        value = find_objects(self._tweet, "metric_type", metric_key, recursive=False, none_value={})
        return value.get('metric_value', 0.0)

    def __repr__(self):
        return "TweetAnalytics(expands={}, engagements={}, follows={}, impressions={}, link_clicks={}, profile_visits={})".format(
            self.expands, self.engagements, self.follows, self.impressions, self.link_clicks, self.profile_visits
        )


class ApiMedia:
    def download(self, filename: str = None, progress_callback: Callable[[str, int, int], None] = None):
        url = self.best_stream().url

        if not url:
            raise ValueError("No Media Download URL found")

        return self._client.http.download_media(url, filename, progress_callback)


class ApiImage(_TwType, ApiMedia):
    def __init__(self, client, media, media_key):
        self._raw = media
        self._client = client
        self.key = media_key
        self.url = self.direct_url = self._raw.get("original_img_url")
        self.width = self._raw.get("original_img_width")
        self.height = self._raw.get("original_img_height")
        self.alt_text = self._raw.get("alt_text")

    def best_stream(self):
        return self

    def __repr__(self):
        return "ApiImage(key={}, width={}, height={}, alt_text={})".format(
            self.key, self.width, self.height, self.alt_text
        )


class ApiVideoVariant(_TwType):
    def __init__(self, client, variant):
        self._client = client
        self._raw = variant
        self.content_type = self._raw.get("content_type")
        self.bit_rate = self._raw.get("bit_rate")
        self.url = self.direct_url = self._raw.get("url")

    def __repr__(self):
        return "ApiVideoVariant(content_type={}, bit_rate={})".format(
            self.content_type, self.bit_rate
        )


class ApiGif(_TwType, ApiMedia):
    def __init__(self, client, media, media_key):
        self._raw = media
        self._client = client
        self._aspect_ratio = self._raw.get("aspect_ratio", {})
        self.key = media_key
        self.preview_image = ApiImage(self._client, self._raw.get("preview_image"), None)
        self.alt_text = self._raw.get("alt_text")
        self.variants = [ApiVideoVariant(self._client, i) for i in self._raw.get("variants", [])]

    def best_stream(self):
        bit_rates = [int(getattr(i, "bit_rate", 0)) for i in self.variants]
        max_bit_rate = max(bit_rates)
        for variant in self.variants:
            if int(getattr(variant, "bit_rate", 0)) == max_bit_rate:
                return variant

        return None

    def __repr__(self):
        return "ApiGif(key={}, variants={})".format(
            self.key, self.variants
        )


class ApiVideo(_TwType, ApiMedia):
    def __init__(self, client, media, media_key):
        self._raw = media
        self._client = client
        self._aspect_ratio = self._raw.get("aspect_ratio", {})
        self.key = media_key
        self.duration_ms = self._raw.get("duration_millis")
        self.alt_text = self._raw.get("alt_text")
        self.preview_image = ApiImage(self._client, self._raw.get("preview_image"), None)
        self.aspect_ratio = f"{self._aspect_ratio.get('numerator')}/{self._aspect_ratio.get('denominator')}"
        self.variants = [ApiVideoVariant(self._client, i) for i in self._raw.get("variants", [])]

    def best_stream(self):
        bit_rates = [int(getattr(i, "bit_rate", 0)) for i in self.variants]
        max_bit_rate = max(bit_rates)
        for variant in self.variants:
            if int(getattr(variant, "bit_rate", 0)) == max_bit_rate:
                return variant

        return None

    def __repr__(self):
        return "ApiVideo(key={}, duration_ms={}, aspect_ratio={}, variants={})".format(
            self.key, self.duration_ms, self.aspect_ratio, self.variants
        )


class ScheduledTweet(_TwType):
    def __init__(self, client, tweet):
        self._raw = tweet
        self._client = client
        self._info = find_objects(self._raw, "scheduling_info", None, none_value={})
        self._create_request = find_objects(self._raw, "tweet_create_request", None, none_value={})
        self._media_entities = find_objects(self._raw, "media_entities", None, none_value=[])
        self.id = self._raw.get("rest_id")
        self.execute_at = self.date = parse_time(self._info.get("execute_at"))
        self.state = self._info.get("state")
        self.text = self._create_request.get("status")
        self.reply_to_tweet_id = self._create_request.get("in_reply_to_status_id")
        self.media = self._parse_media()

    def _parse_media(self):
        medias = []
        for entity in self._media_entities:
            media_typename = entity.get("media_info", {}).get("__typename", "")
            if media_typename == "ApiImage":
                medias.append(ApiImage(self._client, entity["media_info"], entity["media_key"]))
            elif media_typename == "ApiVideo":
                medias.append(ApiVideo(self._client, entity["media_info"], entity["media_key"]))
            elif media_typename == "ApiGif":
                medias.append(ApiGif(self._client, entity["media_info"], entity["media_key"]))
        return medias

    def delete(self):
        return self._client.delete_scheduled_tweet(self.id)

    def __repr__(self):
        return "ScheduledTweet(id={}, date={}, text={}, media={})".format(
            self.id, self.date, self.text, self.media
        )


class Article(_TwType):
    def __init__(self, client, article):
        self._raw = article
        self._client = client
        self._metadata = self._raw.get("metadata", {})
        self._lifecycle = self._raw.get("lifecycle_state", {})
        self.id = self._raw.get("rest_id")
        self.created_at = self.date = parse_time(self._metadata.get("first_published_at_secs"))
        self.edited_at = parse_time(self._lifecycle.get("modified_at_secs"))
        self.title = self._raw.get("title")
        self.preview_text = self._raw.get("preview_text")
        self.cover_media = self._get_media_entity(self._raw.get("cover_media"))
        self.text = self._raw.get("plain_text")
        self.media = [self._get_media_entity(i) for i in self._raw.get("media_entities", [])]

    def _get_media_entity(self, media):
        if not media:
            return None

        parsed_media = None
        _entity_media = media
        media_typename = _entity_media.get("media_info", {}).get("__typename", "")
        if media_typename == "ApiImage":
            parsed_media = ApiImage(self._client, _entity_media["media_info"], _entity_media["media_key"])
        elif media_typename == "ApiVideo":
            parsed_media = ApiVideo(self._client, _entity_media["media_info"], _entity_media["media_key"])
        elif media_typename == "ApiGif":
            parsed_media = ApiGif(self._client, _entity_media["media_info"], _entity_media["media_key"])

        return parsed_media

    def __repr__(self):
        return "Article(id={}, date={}, title={})".format(
            self.id, self.date, self.title
        )
