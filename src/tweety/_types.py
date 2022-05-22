import csv
import sys
import traceback
from dateutil import parser
import openpyxl
try:
    import wget
except ModuleNotFoundError:
    import warnings
    warnings.warn(' "wget" not found in system ,you will not be able to download the medias')
WORKBOOK_HEADERS = ['Created on', 'author', 'is_retweet', 'is_reply', 'tweet_id', 'tweet_body', 'language', 'likes',
                    'retweet_count', 'source', 'medias', 'user_mentioned', 'urls', 'hashtags', 'symbols']


def bar_progress(current, total, width=80):
    progress_message = "Downloading: %d%% [%d / %d] bytes" % (current / total * 100, current, total)
    sys.stdout.write("\r" + progress_message)
    sys.stdout.flush()


class UserTweets:
    def __init__(self, dictionary, user):
        self.dict_ = dictionary
        self.user = user

    def to_xlsx(self, filename=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        for v, i in enumerate(WORKBOOK_HEADERS):
            ws.cell(row=1, column=v + 1).value = i
        max_row = 1
        for p in self.dict_:
            p = p.to_dict()
            ws[f'A{max_row + 1}'] = p['created_on']
            ws[f'B{max_row + 1}'] = p['author'].name
            ws[f'C{max_row + 1}'] = p['is_retweet']
            ws[f'D{max_row + 1}'] = p['is_reply']
            ws[f'E{max_row + 1}'] = p['tweet_id']
            ws[f'F{max_row + 1}'] = p['tweet_body']
            ws[f'G{max_row + 1}'] = p['language']
            ws[f'H{max_row + 1}'] = p['likes']
            ws[f'I{max_row + 1}'] = p['retweet_counts']
            ws[f'J{max_row + 1}'] = p['source']
            media_ = ""
            if not isinstance(p['media'], str) and p['media'] is not None:
                for media in p['media']:
                    media_ = f"{media_} {media.expanded_url};"
                ws[f'K{max_row + 1}'] = media_
            else:
                ws[f'K{max_row + 1}'] = p['media']
            user_mentions = ""
            if not isinstance(p['user_mentions'], str) and p['user_mentions'] is not None:
                for user in p['user_mentions']:
                    user_mentions = f"{user_mentions} {user.screen_name};"
                ws[f'L{max_row + 1}'] = user_mentions
            else:
                ws[f'L{max_row + 1}'] = p['user_mentions']
            urls = ""
            if not isinstance(p['urls'], str):
                for url in p['urls']:
                    urls = f"{urls} {url['expanded_url']};"
                ws[f'M{max_row + 1}'] = urls
            else:
                ws[f'M{max_row + 1}'] = p['urls']
            hashtags = ""
            if not isinstance(p['hashtags'], str):
                for tag in p['hashtags']:
                    hashtags = f"{hashtags} {tag['text']};"
                ws[f'N{max_row + 1}'] = hashtags
            else:
                ws[f'N{max_row + 1}'] = p['hashtags']
            ws[f'O{max_row + 1}'] = p['symbols']
            max_row = max_row + 1
        if not filename:
            filename = f"tweets-{self.user}.xlsx"
        wb.save(filename)

    def to_csv(self, filename=None):
        if not filename:
            filename = f"tweets-{self.user}.csv"
        with open(filename, "w", encoding="ASCII", errors="ignore", newline="") as csv_:
            writer = csv.writer(csv_)
            writer.writerow(WORKBOOK_HEADERS)
            for p in self.dict_:
                p = p.to_dict()
                created_on = p.get('created_at') if p.get('created_at') else p.get('created_on')
                author = p['author'].name
                is_retweet = p['is_retweet']
                is_reply = p['is_reply']
                tweet_id = p['tweet_id']
                tweet_body = p['tweet_body']
                language = p['language']
                likes = p['likes']
                retweet_count = p['retweet_counts']
                source = p['source']
                symbols = p['symbols']
                media_ = ""
                if not isinstance(p['media'], str) and p['media'] is not None:
                    for media in p['media']:
                        media_ = f"{media_} {media.expanded_url};"
                else:
                    media_ = p['media']
                user_mentions = ""
                if not isinstance(p['user_mentions'], str) and p['user_mentions'] is not None:
                    for user in p['user_mentions']:
                        user_mentions = f"{user_mentions} {user.screen_name};"
                else:
                    user_mentions = p['user_mentions']
                hashtags = ""
                if not isinstance(p['hashtags'], str):
                    for tag in p['hashtags']:
                        hashtags = f"{hashtags} {tag['text']};"
                else:
                    hashtags = p['hashtags']
                urls = ""
                if not isinstance(p['urls'], str):
                    for url in p['urls']:
                        urls = f"{urls} {url['expanded_url']};"
                else:
                    urls = p['urls']
                row = [created_on, author, is_retweet, is_reply, tweet_id, tweet_body, language, likes, retweet_count,
                       source, media_, user_mentions, urls, hashtags, symbols]
                writer.writerow(row)

    def to_dict(self):
        return self.dict_

    def __getitem__(self,index):
        return self.dict_[index]

    def __iter__(self):
        for __tweet in self.dict_:
            yield __tweet

    def __repr__(self):
        return f"UserTweets(user={self.user}, count={len(self.dict_)})"


class Tweet:
    def __init__(self, tweet_dict, threads=None):
        self.__dictionary = tweet_dict
        self.id = self.__dictionary.get("tweet_id")
        self.author = self.__dictionary.get("author")
        self.created_on = self.__dictionary.get("created_on")
        self.is_retweet = self.__dictionary.get("is_retweet")
        self.is_reply = self.__dictionary.get("is_reply")
        self.tweet_body = self.__dictionary.get("tweet_body")
        self.language = self.__dictionary.get("language")
        self.likes = self.__dictionary.get("likes")
        self.retweet_counts = self.__dictionary.get("retweet_counts")
        self.card = self.__dictionary.get("card")
        self.media = self.__dictionary.get("media")
        self.user_mentions = self.__dictionary.get("user_mentions")
        self.urls = self.__dictionary.get("urls")
        self.hashtags = self.__dictionary.get("hashtags")
        self.symbols = self.__dictionary.get("symbols")
        self.reply_to = self.__dictionary.get("reply_to")
        self.threads = threads

    def __repr__(self):
        return f"Tweet(id={self.id}, author={self.author}, created_on={self.created_on}, threads={len(self.threads) if self.threads else None})"

    def __iter__(self):
        if self.threads:
            for thread__ in self.threads:
                yield thread__

    def to_dict(self):
        result = self.__dictionary
        result['threads'] = self.threads
        return result


class Media:
    def __init__(self, media_dict):
        self.__dictionary = media_dict
        self.display_url = self.__dictionary.get("display_url")
        self.expanded_url = self.__dictionary.get("expanded_url")
        self.id = self.__dictionary.get("id_str")
        self.indices = self.__dictionary.get("indices")
        self.media_url_https = self.__dictionary.get("media_url_https")
        self.type = self.__dictionary.get("type")
        self.url = self.__dictionary.get("url")
        self.features = self.__dictionary.get("features")
        self.media_key = self.__dictionary.get("media_key")
        self.mediaStats = self.__dictionary.get("mediaStats")
        self.sizes = self.__dictionary.get("sizes")
        self.original_info = self.__dictionary.get("original_info")
        self.file_format = self.media_url_https.split(".")[-1] if self.type == "photo" else None
        self.streams = []
        if self.type == "video":
            self.__parse_video_streams()

    def __parse_video_streams(self):
        videoDict = self.__dictionary.get("video_info")
        if videoDict:
            for i in videoDict.get("variants"):
                if not i.get("content_type").split("/")[-1] == "x-mpegURL":
                    self.streams.append(
                        Stream(
                            i,
                            videoDict.get("duration_millis"),
                            videoDict.get("aspect_ratio"),
                        )
                    )

    def __repr__(self):
        return f"Media(id={self.id}, type={self.type})"

    def download(self, filename_,show_progress=True):
        if show_progress:
            show_progress = bar_progress
        else:
            show_progress = None
        if self.type == "photo":
            filename = f"{filename_}.{self.file_format}"
            wget.download(url=self.media_url_https, out=filename,bar=show_progress)
            if show_progress:
                sys.stdout.write("\n")
            return filename
        elif self.type == "video":
            _res = [int(stream.res) for stream in self.streams if stream.res]
            max_res = max(_res)
            for stream in self.streams:
                if int(stream.res) == max_res:
                    file_format = stream.content_type.split("/")[-1]
                    if not file_format == "x-mpegURL":
                        filename = f"{filename_}.{file_format}"
                        wget.download(url=stream.url, out=filename,bar=show_progress)
                        if show_progress:
                            sys.stdout.write("\n")
                        return filename
        return None

    def to_dict(self):
        return self.__dictionary


class Stream:
    def __init__(self, videoDict, length, ratio):
        self.__dictionary = videoDict
        self.bitrate = self.__dictionary.get("bitrate")
        self.content_type = self.__dictionary.get("content_type")
        self.url = self.__dictionary.get("url")
        self.length = f"{length} millis"
        self.aspect_ratio = ratio
        try:
            self.res = int(self.url.split("/")[7].split("x")[0]) * int(self.url.split("/")[7].split("x")[1])
        except ValueError:
            try:
                self.res = int(self.url.split("/")[6].split("x")[0]) * int(self.url.split("/")[6].split("x")[1])
            except ValueError:
                self.res = None

    def __repr__(self):
        return f"Stream(content_type={self.content_type}, length={self.length}, bitrate={self.bitrate}, res={self.res})"

    def download(self, filename_=None,show_progress=False):
        if show_progress:
            show_progress = bar_progress
        else:
            show_progress = None
        file_format = self.content_type.split("/")[-1]
        if filename_:
            filename = f"{filename_}.{file_format}"
        else:
            filename = None
        wget.download(url=self.url, out=filename,bar=show_progress)
        if show_progress:
            sys.stdout.write("\n")
        return filename


class ShortUser:
    def __init__(self, user_dict):
        self.__dictionary = user_dict
        self.id = self.__dictionary.get("id_str")
        self.name = self.__dictionary.get("name")
        self.screen_name = self.__dictionary.get("screen_name")

    def __repr__(self):
        return f"ShortUser(id={self.id}, name={self.name})>"

    def to_dict(self):
        return self.__dictionary


class User:
    def __init__(self, user_dict, type_=1):
        if type_ == 1:
            self.__dictionary = user_dict['data']['user']['result']
        elif type_ == 2:
            self.__dictionary = user_dict
        else:
            self.__dictionary = user_dict['user_results']['result']
        self.id = self.__dictionary.get("id")
        self.rest_id = self.__dictionary.get("rest_id") if self.__dictionary.get("rest_id") else self.__dictionary.get("id_str")
        self.created_at = parser.parse(self.__dictionary.get("created_at")) if type_ == 2 else parser.parse(self.__dictionary.get("legacy").get("created_at"))
        self.default_profile = self.__dictionary.get("default_profile") if type_ == 2 else self.__dictionary.get("legacy").get("default_profile")
        self.default_profile_image = self.__dictionary.get("default_profile_image") if type_ == 2 else self.__dictionary.get("legacy").get("default_profile_image")
        self.description = self.__dictionary.get("description") if type_ == 2 else self.__dictionary.get("legacy").get("description")
        self.entities = self.__dictionary.get("description") if type_ == 2 else self.__dictionary.get("legacy").get("entities")
        self.fast_followers_count = self.__dictionary.get("fast_followers_count") if type_ == 2 else self.__dictionary.get("legacy").get("fast_followers_count")
        self.favourites_count = self.__dictionary.get("favourites_count")  if type_ == 2 else self.__dictionary.get("legacy").get("favourites_count")
        self.followers_count = self.__dictionary.get("followers_count")  if type_ == 2 else self.__dictionary.get("legacy").get("followers_count")
        self.friends_count = self.__dictionary.get("friends_count") if type_ == 2 else self.__dictionary.get("legacy").get("friends_count")
        self.has_custom_timelines = self.__dictionary.get("has_custom_timelines") if type_ == 2 else self.__dictionary.get("legacy").get("has_custom_timelines")
        self.is_translator = self.__dictionary.get("is_translator") if type_ == 2 else self.__dictionary.get("legacy").get("is_translator")
        self.listed_count = self.__dictionary.get("listed_count") if type_ == 2 else self.__dictionary.get("legacy").get("listed_count")
        self.location = self.__dictionary.get("location") if type_ == 2 else self.__dictionary.get("legacy").get("location")
        self.media_count = self.__dictionary.get("media_count") if type_ == 2 else self.__dictionary.get("legacy").get("media_count")
        self.name = self.__dictionary.get("name") if type_ == 2 else self.__dictionary.get("legacy").get("name")
        self.normal_followers_count = self.__dictionary.get("normal_followers_count") if type_ == 2 else self.__dictionary.get("legacy").get("normal_followers_count")
        self.profile_banner_url = self.__dictionary.get("profile_banner_url") if type_ == 2 else self.__dictionary.get("legacy").get("profile_banner_url")
        self.profile_image_url_https = self.__dictionary.get("profile_image_url_https") if type_ == 2 else self.__dictionary.get("legacy").get("profile_image_url_https")
        self.profile_interstitial_type = self.__dictionary.get("profile_interstitial_type") if type_ == 2 else self.__dictionary.get("legacy").get("profile_interstitial_type")
        self.protected = self.__dictionary.get("protected") if type_ == 2 else self.__dictionary.get("legacy").get("protected")
        self.screen_name = self.__dictionary.get("screen_name") if type_ == 2 else self.__dictionary.get("legacy").get("screen_name")
        self.statuses_count = self.__dictionary.get("statuses_count") if type_ == 2 else self.__dictionary.get("legacy").get("statuses_count")
        self.translator_type = self.__dictionary.get("translator_type") if type_ == 2 else self.__dictionary.get("legacy").get("translator_type")
        self.verified = self.__dictionary.get("verified") if type_ == 2 else self.__dictionary.get("legacy").get("verified")
        self.profile_url = "https://twitter.com/{}".format(self.screen_name)

    def __repr__(self):
        return f"User(id={self.rest_id}, name={self.name}, screen_name={self.screen_name}, followers={self.followers_count}, verified={self.verified})"

    def to_dict(self):
        return self.__dictionary


class Search:
    def __init__(self, dictionary,keyword,filterType=None):
        self.dict_ = dictionary
        self.keyword = keyword
        self.filter = filterType

    def to_xlsx(self, filename=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        for v, i in enumerate(WORKBOOK_HEADERS):
            ws.cell(row=1, column=v + 1).value = i
        max_row = 1
        if not self.filter:
            for p in self.dict_:
                p = p.to_dict()
                ws[f'A{max_row + 1}'] = p['created_on']
                ws[f'B{max_row + 1}'] = p['author'].name
                ws[f'C{max_row + 1}'] = p['is_retweet']
                ws[f'D{max_row + 1}'] = p['is_reply']
                ws[f'E{max_row + 1}'] = p['tweet_id']
                ws[f'F{max_row + 1}'] = p['tweet_body']
                ws[f'G{max_row + 1}'] = p['language']
                ws[f'H{max_row + 1}'] = p['likes']
                ws[f'I{max_row + 1}'] = p['retweet_counts']
                ws[f'J{max_row + 1}'] = p['source']
                media_ = ""
                if not isinstance(p['media'], str) and p['media'] is not None:
                    for media in p['media']:
                        media_ = f"{media_} {media.expanded_url};"
                    ws[f'K{max_row + 1}'] = media_
                else:
                    ws[f'K{max_row + 1}'] = p['media']
                user_mentions = ""
                if not isinstance(p['user_mentions'], str) and p['user_mentions'] is not None:
                    for user in p['user_mentions']:
                        user_mentions = f"{user_mentions} {user.screen_name};"
                    ws[f'L{max_row + 1}'] = user_mentions
                else:
                    ws[f'L{max_row + 1}'] = p['user_mentions']
                urls = ""
                if not isinstance(p['urls'], str):
                    for url in p['urls']:
                        urls = f"{urls} {url['expanded_url']};"
                    ws[f'M{max_row + 1}'] = urls
                else:
                    ws[f'M{max_row + 1}'] = p['urls']
                hashtags = ""
                if not isinstance(p['hashtags'], str):
                    for tag in p['hashtags']:
                        hashtags = f"{hashtags} {tag['text']};"
                    ws[f'N{max_row + 1}'] = hashtags
                else:
                    ws[f'N{max_row + 1}'] = p['hashtags']
                ws[f'O{max_row + 1}'] = p['symbols']
                max_row = max_row + 1
        else:
            print("to_xlsx() method is not yet supported for only users search filter")
        if not filename:
            filename = f"searches-{self.keyword}.xlsx"
        wb.save(filename)

    def to_csv(self, filename=None):
        if not filename:
            filename = f"searches-{self.keyword}.csv"
        if not self.filter:
            with open(filename, "w", encoding="ASCII", errors="ignore", newline="") as csv_:
                writer = csv.writer(csv_)
                writer.writerow(WORKBOOK_HEADERS)
                for p in self.dict_:
                    p = p.to_dict()
                    created_on = p.get('created_at') if p.get('created_at') else p.get('created_on')
                    author = p['author'].name
                    is_retweet = p['is_retweet']
                    is_reply = p['is_reply']
                    tweet_id = p['tweet_id']
                    tweet_body = p['tweet_body']
                    language = p['language']
                    likes = p['likes']
                    retweet_count = p['retweet_counts']
                    source = p['source']
                    symbols = p['symbols']
                    media_ = ""
                    if not isinstance(p['media'], str) and p['media'] is not None:
                        for media in p['media']:
                            media_ = f"{media_} {media.expanded_url};"
                    else:
                        media_ = p['media']
                    user_mentions = ""
                    if not isinstance(p['user_mentions'], str) and p['user_mentions'] is not None:
                        for user in p['user_mentions']:
                            user_mentions = f"{user_mentions} {user.screen_name};"
                    else:
                        user_mentions = p['user_mentions']
                    hashtags = ""
                    if not isinstance(p['hashtags'], str):
                        for tag in p['hashtags']:
                            hashtags = f"{hashtags} {tag['text']};"
                    else:
                        hashtags = p['hashtags']
                    urls = ""
                    if not isinstance(p['urls'], str):
                        for url in p['urls']:
                            urls = f"{urls} {url['expanded_url']};"
                    else:
                        urls = p['urls']
                    row = [created_on, author, is_retweet, is_reply, tweet_id, tweet_body, language, likes, retweet_count,
                           source, media_, user_mentions, urls, hashtags, symbols]
                    writer.writerow(row)
        else:
            print("to_csv() method is not yet supported for only users search filter")

    def to_dict(self):
        return self.dict_

    def __getitem__(self,index):
        return self.dict_[index]

    def __iter__(self):
        for __tweet in self.dict_:
            yield __tweet

    def __repr__(self):
        return f"Search(keyword={self.keyword} , count={len(self.dict_)})"


class Trends:
    def __init__(self, trends_dict):
        self.__dictionary = trends_dict
        self.name = self.__dictionary.get("name")
        self.url = self.__dictionary.get("url")
        self.tweet_count = self.__dictionary.get("tweet_count")

    def __repr__(self):
        return f"Trends(name={self.name})"

    def to_dict(self):
        return self.__dictionary


class UserLegacy:
    def __init__(self, user_dict):
        self.__dictionary = user_dict
        self.id = self.__dictionary.get("id")
        self.rest_id = self.__dictionary.get("id")
        self.created_at = parser.parse(self.__dictionary.get("created_at")) if self.__dictionary.get("created_at") else None
        self.default_profile = self.__dictionary.get("default_profile")
        self.default_profile_image = self.__dictionary.get("default_profile_image")
        self.description = self.__dictionary.get("description")
        self.entities = self.__dictionary.get("entities")
        self.fast_followers_count = self.__dictionary.get("fast_followers_count")
        self.favourites_count = self.__dictionary.get("favourites_count")
        self.followers_count = self.__dictionary.get("followers_count")
        self.friends_count = self.__dictionary.get("friends_count")
        self.has_custom_timelines = self.__dictionary.get("has_custom_timelines")
        self.is_translator = self.__dictionary.get("is_translator")
        self.listed_count = self.__dictionary.get("listed_count")
        self.location = self.__dictionary.get("location")
        self.media_count = self.__dictionary.get("media_count")
        self.name = self.__dictionary.get("name")
        self.normal_followers_count = self.__dictionary.get("normal_followers_count")
        self.profile_banner_url = self.__dictionary.get("profile_banner_url")
        self.profile_image_url_https = self.__dictionary.get("profile_image_url_https")
        self.profile_interstitial_type = self.__dictionary.get("profile_interstitial_type")
        self.protected = self.__dictionary.get("protected")
        self.screen_name = self.__dictionary.get("screen_name")
        self.statuses_count = self.__dictionary.get("statuses_count")
        self.translator_type = self.__dictionary.get("translator_type")
        self.verified = self.__dictionary.get("verified")

    def __repr__(self):
        return f"User(id={self.rest_id}, name={self.name}, followers={self.followers_count} , verified={self.verified})"

    def to_dict(self):
        return self.__dictionary


class Card:
    def __init__(self,card_dict):
        self._dict = card_dict
        self.__bindings = self._dict['legacy'].get("binding_values")
        self.rest_id = self._dict.get("rest_id")
        self.name = self._dict['legacy'].get("name")
        self.choices = []
        self.end_time = None
        self.last_updated_time = None
        self.duration = None
        self.user_ref = self._dict['legacy'].get("user_refs")
        self.__parse_choices()

    def __parse_choices(self):
        for _ in self.__bindings:
            _key = _.get("key").split("_")
            if "choice" in _key[0] and "label" in _key[1]:
                _cardName = _key[0]
                _cardValue = _['value']['string_value']
                _cardValueType = _['value']['type']
                _cardCounts = 0
                _cardCountsType = None
                for __ in self.__bindings:
                    __key = __.get("key")
                    if __key[0] == _key[0] and "count" in __key[1]:
                        _cardCounts = __['value']['string_value']
                        _cardCountsType = __['value']['type']
                _r = {
                    "card_name":_cardName,
                    "card_value":_cardValue,
                    "card_value_type":_cardValueType,
                    "card_counts":_cardCounts,
                    "card_counts_type":_cardCountsType,
                }
                self.choices.append(Choice(_r))
            elif _key[0] == "end" and _key[1] == "datetime":
                self.end_time = parser.parse(_['value']['string_value'])
                # last_updated_datetime_utc
            elif _key[0] == "last" and _key[1] == "updated":
                self.last_updated_time = parser.parse(_['value']['string_value'])
                # duration_minutes
            elif _key[0] == "duration" and _key[1] == "minutes":
                self.duration = _['value']['string_value']

    def __repr__(self):
        return f"Card(id={self.rest_id}, choices={len(self.choices) if self.choices else []}, duration={len(self.duration)})"


class Choice:
    def __init__(self,_dict):
        self._dict = _dict
        self.name = self._dict.get("card_name")
        self.value = self._dict.get("card_value")
        self.type = self._dict.get("card_value_type")
        self.counts = self._dict.get("card_counts")
        self.counts_type = self._dict.get("card_counts_type")

    def __repr__(self):
        return f"Choice(name={self.name}, value={self.value}, counts={self.counts})"
